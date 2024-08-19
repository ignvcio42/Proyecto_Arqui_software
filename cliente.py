import socket
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
import os
from datetime import datetime

# Create a TCP/IP socket
sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

# Connect the socket to the port where the bus is listening
bus_address = ('localhost', 5000)
print('connecting to {} port {}'.format(*bus_address))
sock.connect(bus_address)

def send_message(service, action, data):
    message_data = f"{action}{data}"
    message = f"{len(message_data):05}{service}{message_data}".encode()
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    bus_address = ('localhost', 5000)
    sock.connect(bus_address)
    try:
        sock.sendall(message)
        amount_expected = int(sock.recv(5))
        amount_received = 0
        response = b''
        while amount_received < amount_expected:
            chunk = sock.recv(amount_expected - amount_received)
            amount_received += len(chunk)
            response += chunk
    finally:
        sock.close()
    return response.decode()

def validar_rut(rut):
    if rut.isdigit() and len(rut) <= 9:
        return True
    return False

def validar_id(id):
    if id.isdigit():
        return True
    return False

# Funciones para la gestión de equipos
def obtener_info_equipo(id_equipo):
    if not validar_id(id_equipo):
        print("Error: El ID debe ser un número entero.")
        return
    response = send_message("EQUIP", "CODIU", str(id_equipo))
    response_parts = response.split(',')
    if response_parts[0] == "EQUIPOKOK":
        id = response_parts[1]
        nombre = response_parts[2]
        descripcion = response_parts[3]
        tipo = response_parts[4]
        tarifa = response_parts[5]
        arrendado = response_parts[6]

        print(f"ID: {id}")
        print(f"Nombre: {nombre}")
        print(f"Descripción: {descripcion}")
        print(f"Tipo: {tipo}")
        print(f"Tarifa: {tarifa}")
        if arrendado != "No arrendado":
            print(f"Arrendado: {arrendado}")
    else:
        print("Error al obtener la información del equipo.")

def obtener_info_todos_equipos():
    response = send_message("EQUIP", "CODIT", "")
    if "EQUIPOK" in response:
        response = response[10:]
        partes = response.split(';')
        if len(partes) != 2:
            print("Formato de respuesta inválido.")
            return

        disponibles = partes[0].replace("Disponibles:", "").split('|')
        arrendados = partes[1].replace("Arrendados:", "").split('|')

        print("Dispositivos Disponibles:")
        for equipo in disponibles:
            if equipo:
                response_parts = equipo.split(',')
                if len(response_parts) >= 5:
                    id = response_parts[0]
                    nombre = response_parts[1]
                    descripcion = response_parts[2]
                    tipo = response_parts[3]
                    tarifa = response_parts[4]
                    print(f"ID: {id}")
                    print(f"Nombre: {nombre}")
                    print(f"Descripción: {descripcion}")
                    print(f"Tipo: {tipo}")
                    print(f"Tarifa: {tarifa}")
                    print("-----------------")

        print("Dispositivos Arrendados:")
        for equipo in arrendados:
            if equipo:
                response_parts = equipo.split(',')
                if len(response_parts) >= 5:
                    id = response_parts[0]
                    nombre = response_parts[1]
                    descripcion = response_parts[2]
                    tipo = response_parts[3]
                    tarifa = response_parts[4]
                    print(f"ID: {id}")
                    print(f"Nombre: {nombre}")
                    print(f"Descripción: {descripcion}")
                    print(f"Tipo: {tipo}")
                    print(f"Tarifa: {tarifa}")
                    print("-----------------")
    else:
        print("Error al obtener la información de los equipos.")

def añadir_equipo(nombre, descripcion, tipo, tarifa):
    data = f"{nombre},{descripcion},{tipo},{tarifa}"
    response = send_message("EQUIP", "CODAE", data)
    if response == "EQUIPOKNK, Acción inválida":
        print("Error al añadir equipo")
    else:
        print(f"Equipo: {nombre} ha sido añadido")

def eliminar_equipo(id_equipo):
    if not validar_id(id_equipo):
        print("Error: El ID debe ser un número entero.")
        return
    response = send_message("EQUIP", "CODEE", str(id_equipo))
    if response == "EQUIPOKNK, Acción inválida":
        print("Error al eliminar equipo")
    else:
        print(f"Equipo de ID: {id_equipo} ha sido eliminado")

def modificar_equipo(id_equipo, nombre, descripcion, tipo, tarifa):
    if not validar_id(id_equipo):
        print("Error: El ID debe ser un número entero.")
        return
    data = f"{id_equipo},{nombre},{descripcion},{tipo},{tarifa}"
    response = send_message("EQUIP", "CODME", data)
    if response == "EQUIPOKNK, Acción inválida":
        print("Error al modificar equipo")
    else:
        print(f"Datos del equipo de ID: {id_equipo} han sido modificados")

def obtener_dispositivos_disponibles():
    response = send_message("EQUIP", "DISPO", "disponibles")
    dispositivos = response.split(',')
    for dispositivo in dispositivos:
        print(dispositivo)

def obtener_dispositivos_no_disponibles():
    response = send_message("EQUIP", "DISPO", "no_disponibles")
    dispositivos = response.split(',')
    for dispositivo in dispositivos:
        print(dispositivo)    

# Funciones para la gestión de usuarios
def añadir_usuario(nombre, rut, email):
    if not validar_rut(rut):
        print("Error: El RUT debe ser un número entero menor a 9 dígitos.")
        return
    
    data = f"{nombre},{rut},{email}"
    response = send_message("USUAR", "CODAU", data)
    if response == "USUAROKNK, Acción inválida":
        print("Error al añadir usuario")
    else:
        print(f"Usuario de RUT: {rut} ha sido añadido")

def eliminar_usuario(rut):
    if not validar_rut(rut):
        print("Error: El RUT debe ser un número entero menor a 9 dígitos.")
        return

    response = send_message("USUAR", "CODEU", str(rut))
    if response == "USUAROKNK, Acción inválida":
        print("Error al eliminar usuario")
    else:
        print(f"Usuario de RUT: {rut} ha sido eliminado")
    

def modificar_usuario():
    rut = input("Ingrese el RUT del usuario a modificar: ")
    if not validar_rut(rut):
        print("Error: El RUT debe ser un número entero menor a 9 dígitos.")
        return
    
    nombre = input("Ingrese el nuevo nombre del usuario: ")
    email = input("Ingrese el nuevo email del usuario: ")
    data = f"{rut},{nombre},{email}"
    response = send_message("USUAR", "CODMU", data)
    if response == "USUAROKNK, Acción inválida":
        print("Error al modificar usuario")
    else:
        print(f"Datos del usuario de RUT: {rut} han sido modificados")
    

def obtener_info_usuario(rut):
    if not validar_rut(rut):
        print("Error: El RUT debe ser un número entero menor a 9 dígitos.")
        return
    
    response = send_message("USUAR", "CODIU", str(rut))
    response_parts = response.split(',')
    
    if response_parts[0] == "USUAROKOK":
        nombre = response_parts[1]
        rut = response_parts[2]
        email = response_parts[3]
        print(f"Nombre: {nombre}")
        print(f"RUT: {rut}")
        print(f"Email: {email}")
    else:
        print("Error al obtener la información del usuario.")

def obtener_info_todos_usuarios():
    response = send_message("USUAR", "CODIT", "")
    if "USUAROK" in response:
        response = response[10:]
        usuarios = response.split('|')
        if not usuarios or usuarios == ['']:
            print("No hay usuarios en la base de datos.")
            return
        for usuario in usuarios:
            response_parts = usuario.split(',')
            if len(response_parts) >= 3:
                nombre = response_parts[0]
                rut = response_parts[1]
                email = response_parts[2]
                print(f"Nombre: {nombre}")
                print(f"RUT: {rut}")
                print(f"Email: {email}")
                print("-----------------")
    else:
        print("Error al obtener la información de los usuarios.")

# Funciones para la gestión de alimentos
def añadir_alimento(nombre, precio, stock):
    data = f"{nombre},{precio},{stock}"
    response = send_message("ALIME", "CODAA", data)
    if response == "ALIMEOKNK, Acción inválida":
        print("Error al añadir alimento")
    else: 
        print(f"Alimento: {nombre} ha sido añadido")

def eliminar_alimento(id_alimento):
    if not validar_id(id_alimento):
        print("Error: El ID debe ser un número entero.")
        return
    response = send_message("ALIME", "CODEA", str(id_alimento))
    if response == "ALIMEOKNK, Acción inválida":
        print("Error al eliminar alimento")
    else:
        print(f"Alimento de ID: {id_alimento} ha sido eliminado")

def modificar_alimento(id_alimento, nombre, precio, stock):
    if not validar_id(id_alimento):
        print("Error: El ID debe ser un número entero.")
        return
    data = f"{id_alimento},{nombre},{precio},{stock}"
    response = send_message("ALIME", "CODMA", data)
    if response == "ALIMEOKNK, Acción inválida":
        print("Error al modificar alimento")
    else:
        print(f"Los datos del alimento de ID: {id_alimento} han sido modificados")

def obtener_info_alimento(id_alimento):
    if not validar_id(id_alimento):
        print("Error: El ID debe ser un número entero.")
        return
    response = send_message("ALIME", "CODIA", str(id_alimento))
    response_parts = response.split(',')
    
    if response_parts[0] == "ALIMEOKOK":
        id = response_parts[1]
        nombre = response_parts[2]
        precio = response_parts[3]
        stock = response_parts[4]
        print(f"ID: {id}")
        print(f"Nombre: {nombre}")
        print(f"Precio: {precio}")
        print(f"Stock: {stock}")
    else:
        print("Error al obtener la información del alimento.")

def obtener_info_todos_alimentos():
    response = send_message("ALIME", "CODIT", "")
    if "ALIMEOK" in response:
        response = response[10:]
        alimentos = response.split('|')
        if not alimentos or alimentos == ['']:
            print("No hay alimentos en la base de datos.")
            return
        for alimento in alimentos:
            response_parts = alimento.split(',')
            if len(response_parts) >= 4:
                id = response_parts[0]
                nombre = response_parts[1]
                precio = response_parts[2]
                stock = response_parts[3]
                print(f"ID: {id}")
                print(f"Nombre: {nombre}")
                print(f"Precio: {precio}")
                print(f"Stock: {stock}")
                print("-----------------")
    else:
        print("Error al obtener la información de los alimentos.")

# Funciones para la gestión de arriendo de equipos
def arrendar_equipo(rut_usuario, id_equipo, tiempo_arriendo):
    if not validar_rut(rut_usuario):
        print("Error: El RUT debe ser un número entero menor a 9 dígitos.")
        return
    if not validar_id(id_equipo):
        print("Error: El ID del equipo debe ser un número entero.")
        return
    if not tiempo_arriendo.isdigit():
        print("Error: El tiempo de arriendo debe ser un número entero.")
        return
    
    data = f"{rut_usuario},{id_equipo},{tiempo_arriendo}"
    response = send_message("ARRIE", "CODAE", data)
    response_parts = response.split(',')
    if response_parts[0] == "ARRIEOKOK":
        fecha = response_parts[1]
        monto = response_parts[2]
        fecha_fin = response_parts[3]
        print(f"¡Arriendo exitoso! Usuario: {rut_usuario} - Equipo: {id_equipo}")
        print(f"Fecha inicio: {fecha} - Fecha final: {fecha_fin} - Monto: {monto}")
    else:
        print("Error: ", response_parts)

# Funciones para la venta de alimentos
def vender_alimento(rut_usuario, nombre_alimento, cantidad):
    if not validar_rut(rut_usuario):
        print("Error: El RUT debe ser un número entero menor a 9 dígitos.")
        return
    
    data = f"{rut_usuario},{nombre_alimento},{cantidad}"
    response = send_message("VENAL", "CODAC", data)
    response_parts = response.split(',')
    if response_parts[0] == "VENALOKOK":
        total = response_parts[1]
        print("Venta exitosa:")
        print(f"-Usuario: {rut_usuario}")
        print(f"-Producto: {nombre_alimento}")
        print(f"-Cantidad: {cantidad}")      
        print(f"-Monto total: {total}")
    else:
        print("Error: ", response_parts[1])

# Funciones para el registro de ganancias
def obtener_ganancias_arriendo(fecha_inicio, fecha_fin):
    data = f"{fecha_inicio},{fecha_fin}"
    response = send_message("REGAN", "CODAE", data)
    response_parts = response.split(',', 1)
    if response_parts[0] == "REGANOKOK":
        ganancias = response_parts[1].split('|')
        for ganancia in ganancias:
            if ',' in ganancia:
                fecha, monto = ganancia.split(',')
                print(f"Fecha: {fecha}, Monto: {monto}")
            else:
                print(f"Formato incorrecto: {ganancia}")
    else:
        print("Error: ", response_parts[1])

def obtener_ganancias_ventas(fecha_inicio, fecha_fin):
    data = f"{fecha_inicio},{fecha_fin}"
    response = send_message("REGAN", "CODVA", data)
    response_parts = response.split(',', 1)
    if response_parts[0] == "REGANOKOK":
        ganancias = response_parts[1].split('|')
        for ganancia in ganancias:
            if ',' in ganancia:
                fecha, monto = ganancia.split(',')
                print(f"Fecha: {fecha}, Monto: {monto}")
            else:
                print(f"Formato incorrecto: {ganancia}")
    else:
        print("Error: ", response_parts[1])

# Directorio base donde se guardarán los informes
base_directory = "informes"

# Funciones para generar informes en Excel
def generar_excel_ganancia_equipos():
    response = send_message("INFOR", "CODGG", "")
    response_parts = response.split(',', 1)
    if response_parts[0] == "INFOROKOK":
        montos_por_tipo = response_parts[1].split('|')
        
        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        folder_path = os.path.join(base_directory, fecha_actual)
        os.makedirs(folder_path, exist_ok=True)
        
        file_name = f"ganancia_equipos.xlsx"
        file_path = os.path.join(folder_path, file_name)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ganancia Equipos"
        
        header_fill = PatternFill(start_color="0072C6", end_color="0072C6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        cell_alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append(["Tipo de Equipo", "Monto"])
        for tipo_monto in montos_por_tipo:
            subpartes = tipo_monto.split(',')
            ws.append([subpartes[0], float(subpartes[1])])  
            
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = cell_alignment
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        wb.save(file_path)
        print(f"Informe de ganancia por tipos de equipos generado: {file_path}")
    else:
        print("Error al generar el informe: ", response_parts[1])

def generar_excel_uso_equipos():
    try:
        response = send_message("INFOR", "CODGU", "")
        response_parts = response.split(',', 1)
        if response_parts[0] == "INFOROKOK":
            uso_equipos = response_parts[1].split('|')
            
            fecha_actual = datetime.now().strftime("%Y-%m-%d")
            folder_path = os.path.join(base_directory, fecha_actual)
            os.makedirs(folder_path, exist_ok=True)
            
            file_name = f"uso_equipos.xlsx"
            file_path = os.path.join(folder_path, file_name)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Uso Equipos"
            
            header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            cell_alignment = Alignment(horizontal="center", vertical="center")
            
            ws.append(["ID Equipo", "Nombre", "Tiempo de Uso", "Monto Total"])
            for equipo in uso_equipos:
                id_equipo, nombre, tiempo, monto = equipo.split(',')
                ws.append([id_equipo, nombre, float(tiempo), float(monto)])  
            
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = cell_alignment
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            
            wb.save(file_path)
            print(f"Informe de uso de equipos generado: {file_path}")
        else:
            print("Error al generar el informe: ", response_parts[1])
    except Exception as e:
        print(f"Error: {e}")

def generar_excel_ventas_alimentos():
    try:
        response = send_message("INFOR", "CODGV", "")
        response_parts = response.split(',', 1)
        if response_parts[0] == "INFOROKOK":
            ventas_alimentos = response_parts[1].split('|')
            
            fecha_actual = datetime.now().strftime("%Y-%m-%d")
            folder_path = os.path.join(base_directory, fecha_actual)
            os.makedirs(folder_path, exist_ok=True)
            
            file_name = f"ventas_alimentos.xlsx"
            file_path = os.path.join(folder_path, file_name)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ventas Alimentos"
            
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            cell_alignment = Alignment(horizontal="center", vertical="center")
            
            ws.append(["ID Alimento", "Nombre", "Monto"])
            for alimento in ventas_alimentos:
                subpartes = alimento.split(',')
                ws.append([subpartes[0], subpartes[1], float(subpartes[2])])  
            
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = cell_alignment
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            
            wb.save(file_path)
            print(f"Informe de ventas de alimentos generado: {file_path}")
        else:
            print("Error al generar el informe: ", response_parts[1])
    except Exception as e:
        print(f"Error: {e}")

# Funciones para la gestión de juegos
def añadir_juego(nombre, descripcion, id_equipo):
    if not validar_id(id_equipo):
        print("Error: El ID del equipo debe ser un número entero.")
        return
    data = f"{nombre},{descripcion},{id_equipo}"
    response = send_message("JUEGO", "CODAJ", data)
    if response == "JUEGOOKNK, Acción inválida":
        print("Error al añadir juego")
    else:
        print(f"Juego: {nombre} ha sido añadido con ID: {response.split(',')[1]}")

def eliminar_juego(id_juego):
    if not validar_id(id_juego):
        print("Error: El ID del juego debe ser un número entero.")
        return
    response = send_message("JUEGO", "CODEJ", str(id_juego))
    if response == "JUEGOOKNK, Acción inválida":
        print("Error al eliminar juego")
    else:
        print(f"Juego de ID: {id_juego} ha sido eliminado")

def modificar_juego(id_juego, nombre, descripcion, id_equipo):
    if not validar_id(id_juego):
        print("Error: El ID del juego debe ser un número entero.")
        return
    if not validar_id(id_equipo):
        print("Error: El ID del equipo debe ser un número entero.")
        return
    data = f"{id_juego},{nombre},{descripcion},{id_equipo}"
    response = send_message("JUEGO", "CODMJ", data)
    if response == "JUEGOOKNK, Acción inválida":
        print("Error al modificar juego")
    else:
        print(f"Datos del juego de ID: {id_juego} han sido modificados")

def obtener_info_juego(id_juego):
    if not validar_id(id_juego):
        print("Error: El ID del juego debe ser un número entero.")
        return
    response = send_message("JUEGO", "CODIU", str(id_juego))
    response_parts = response.split(',')
    if response_parts[0] == "JUEGOOKOK":
        id_juego = response_parts[1]
        nombre = response_parts[2]
        descripcion = response_parts[3]
        id_equipo = response_parts[4]
        print(f"ID: {id_juego}")
        print(f"Nombre: {nombre}")
        print(f"Descripción: {descripcion}")
        print(f"ID Equipo: {id_equipo}")
    else:
        print("Juego no encontrado")

def obtener_info_todos_juegos():
    response = send_message("JUEGO", "CODIT", "")
    if "JUEGOOK" in response:
        response = response[10:]
        juegos = response.split('|')
        if not juegos or juegos == ['']:
            print("No hay juegos en la base de datos.")
            return
        for juego in juegos:
            response_parts = juego.split(',')
            if len(response_parts) >= 4:
                id_juego = response_parts[0]
                nombre = response_parts[1]
                descripcion = response_parts[2]
                id_equipo = response_parts[3]
                print(f"ID: {id_juego}")
                print(f"Nombre: {nombre}")
                print(f"Descripción: {descripcion}")
                print(f"ID Equipo: {id_equipo}")
                print("-----------------")
    else:
        print("Error al obtener la información de los juegos.")

try:
    while True:
        print("Menú de opciones:")
        print("1. Gestión de equipos")
        print("2. Gestión de usuarios")
        print("3. Gestión de alimentos")
        print("4. Gestión de juegos")
        print("5. Arriendo de equipos")
        print("6. Venta de alimentos")
        print("7. Registro de ganancias")
        print("8. Informes")
        print("9. Salir")
        option = input("Seleccione una opción: ")
        
        if option == "1":
            print("Menú de gestión de equipos:")
            print("1. Añadir equipo")
            print("2. Eliminar equipo")
            print("3. Modificar equipo")
            print("4. Obtener información de un equipo")
            print("5. Obtener información de todos los equipos")
            print("6. Volver al menú principal")
            equip_option = input("Seleccione una opción: ")
            if equip_option == "1":
                nombre = input("Ingrese nombre del equipo: ")
                descripcion = input("Ingrese descripción del equipo: ")
                tipo = input("Ingrese tipo del equipo: ")
                tarifa = input("Ingrese tarifa del equipo: ")
                añadir_equipo(nombre, descripcion, tipo, tarifa)
            elif equip_option == "2":
                id_equipo = input("Ingrese ID del equipo: ")
                eliminar_equipo(id_equipo)
            elif equip_option == "3":
                id_equipo = input("Ingrese ID del equipo: ")
                nombre = input("Ingrese nuevo nombre del equipo: ")
                descripcion = input("Ingrese nueva descripción del equipo: ")
                tipo = input("Ingrese nuevo tipo del equipo: ")
                tarifa = input("Ingrese nueva tarifa del equipo: ")
                modificar_equipo(id_equipo, nombre, descripcion, tipo, tarifa)
            elif equip_option == "4":
                id_equipo = input("Ingrese ID del equipo: ")
                obtener_info_equipo(id_equipo)
            elif equip_option == "5":
                obtener_info_todos_equipos()
            elif equip_option == "6":
                continue
            else:
                print("Opción no válida")
        
        elif option == "2":
            print("Menú de gestión de usuarios:")
            print("1. Añadir usuario")
            print("2. Eliminar usuario")
            print("3. Modificar usuario")
            print("4. Obtener información de un usuario")
            print("5. Obtener información de todos los usuarios")
            print("6. Volver al menú principal")
            
            while True:
                usuar_option = input("Seleccione una opción: ")
                
                if usuar_option == "1":
                    nombre = input("Ingrese nombre del usuario: ")
                    rut = input("Ingrese RUT del usuario: ")
                    email = input("Ingrese email del usuario: ")
                    añadir_usuario(nombre, rut, email)
                    break
                
                elif usuar_option == "2":
                    rut = input("Ingrese RUT del usuario: ")
                    eliminar_usuario(rut)
                    break
                
                elif usuar_option == "3":
                    modificar_usuario()
                    break
                
                elif usuar_option == "4":
                    rut = input("Ingrese RUT del usuario: ")
                    obtener_info_usuario(rut)
                    break
                
                elif usuar_option == "5":
                    obtener_info_todos_usuarios()
                    break
                
                elif usuar_option == "6":
                    break
                
                else:
                    print("Opción no válida")
                    break

        
        elif option == "3":
            print("Menú de gestión de alimentos:")
            print("1. Añadir alimento")
            print("2. Eliminar alimento")
            print("3. Modificar alimento")
            print("4. Obtener información de un alimento")
            print("5. Obtener información de todos los alimentos")
            print("6. Volver al menú principal")
            alime_option = input("Seleccione una opción: ")
            if alime_option == "1":
                nombre = input("Ingrese nombre del alimento: ")
                precio = input("Ingrese precio del alimento: ")
                stock = input("Ingrese stock del alimento: ")
                añadir_alimento(nombre, precio, stock)
            elif alime_option == "2":
                id_alimento = input("Ingrese ID del alimento: ")
                eliminar_alimento(id_alimento)
            elif alime_option == "3":
                id_alimento = input("Ingrese ID del alimento: ")
                nombre = input("Ingrese nombre del alimento: ")
                precio = input("Ingrese precio del alimento: ")
                stock = input("Ingrese stock del alimento: ")
                modificar_alimento(id_alimento, nombre, precio, stock)
            elif alime_option == "4":
                id_alimento = input("Ingrese ID del alimento: ")
                obtener_info_alimento(id_alimento)
            elif alime_option == "5":
                obtener_info_todos_alimentos()
            elif alime_option == "6":
                continue
            else:
                print("Opción no válida")
        elif option == "4":
            print("Menú de gestión de juegos:")
            print("1. Añadir juego")
            print("2. Eliminar juego")
            print("3. Modificar juego")
            print("4. Obtener información de un juego")
            print("5. Obtener información de todos los juegos")
            print("6. Volver al menú principal")
            juego_option = input("Seleccione una opción: ")
            if juego_option == "1":
                nombre = input("Ingrese nombre del juego: ")
                descripcion = input("Ingrese descripción del juego: ")
                id_equipo = input("Ingrese ID del equipo asociado al juego: ")
                añadir_juego(nombre, descripcion, id_equipo)
            elif juego_option == "2":
                id_juego = input("Ingrese ID del juego: ")
                eliminar_juego(id_juego)
            elif juego_option == "3":
                id_juego = input("Ingrese ID del juego: ")
                nombre = input("Ingrese nuevo nombre del juego: ")
                descripcion = input("Ingrese nueva descripción del juego: ")
                id_equipo = input("Ingrese nuevo ID del equipo asociado al juego: ")
                modificar_juego(id_juego, nombre, descripcion, id_equipo)
            elif juego_option == "4":
                id_juego = input("Ingrese ID del juego: ")
                obtener_info_juego(id_juego)
            elif juego_option == "5":
                obtener_info_todos_juegos()
            elif juego_option == "6":
                continue
            else:
                print("Opción no válida")
        elif option == "5":
            print("Menú de arriendo de equipos:")
            print("1. Arrendar equipo")
            print("2. Volver al menú principal")
            arrie_option = input("Seleccione una opción: ")
            if arrie_option == "1":
                rut_usuario = input("Ingrese RUT del usuario: ")
                id_equipo = input("Ingrese ID del equipo: ")
                tiempo_arriendo = input("Ingrese tiempo de arriendo en horas: ")
                arrendar_equipo(rut_usuario, id_equipo, tiempo_arriendo)
            elif arrie_option == "2":
                continue
            else:
                print("Opción no válida")
                
        elif option == "6":
            print("Menú de venta de alimentos:")
            print("1. Vender alimento")
            print("2. Volver al menú principal")
            venal_option = input("Seleccione una opción: ")
            if venal_option == "1":
                rut_usuario = input("Ingrese RUT del usuario: ")
                nombre_alimento = input("Ingrese nombre del alimento: ")
                cantidad = input("Ingrese cantidad del alimento: ")
                vender_alimento(rut_usuario, nombre_alimento, cantidad)
            elif venal_option == "2":
                continue
            else:
                print("Opción no válida")

                
        elif option == "7":
            print("Menú de registro de ganancias:")
            print("1. Obtener registro de ganancias por arriendo de equipos")
            print("2. Obtener registro de ganancias por venta de alimentos")
            print("3. Volver al menú principal")
            regan_option = input("Seleccione una opción: ")
            if regan_option == "1":
                fecha_inicio = input("Ingrese la fecha de inicio (YYYY-MM-DD): ")
                fecha_fin = input("Ingrese la fecha de fin (YYYY-MM-DD): ")
                obtener_ganancias_arriendo(fecha_inicio, fecha_fin)
            elif regan_option == "2":
                fecha_inicio = input("Ingrese la fecha de inicio (YYYY-MM-DD): ")
                fecha_fin = input("Ingrese la fecha de fin (YYYY-MM-DD): ")
                obtener_ganancias_ventas(fecha_inicio, fecha_fin)
            elif regan_option == "3":
                continue
            else:
                print("Opción no válida")
                
        elif option == "8":
            print("Menú de informes:")
            print("1. Generar informe de ganancia de equipos (Excel)")
            print("2. Generar informe de uso de equipos (Excel)")
            print("3. Generar informe de ventas de alimentos (Excel)")
            print("4. Volver al menú principal")
            informe_option = input("Seleccione una opción: ")
            if informe_option == "1":
                generar_excel_ganancia_equipos()
            elif informe_option == "2":
                generar_excel_uso_equipos()
            elif informe_option == "3":
                generar_excel_ventas_alimentos()
            elif informe_option == "4":
                continue
            else:
                print("Opción no válida")
                     
        elif option == "9":
            break
        
        else:
            print("Opción no válida")
finally:
    print('Cerrando el cliente')
    sock.close()
